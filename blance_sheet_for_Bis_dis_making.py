import pandas
import numpy as np
import openpyxl
from openpyxl import load_workbook
import os
def sortbylist(dataindex,datasort):
    
    var34=np.array(dataindex)
    unvar3=np.unique(var34)
    itr=np.size(unvar3)
    k=np.size(var34)
    
    datalist=np.zeros(itr)
    n=0
    while n<itr :
        k2=0
        while k2<k :
            if var34[k2]==unvar3[n]:
                datalist[n]=datalist[n]+datasort[k2]
                k2=k2+1
            else:
                k2=k2+1
        
        
            
        n=n+1
    exp=np.stack((unvar3,datalist),axis=0)
    return exp
def fill_array(A,name,sheet_n,in1):
    name2=name
    name2.replace('"', "'")
    
  
    
    if os.path.exists(name2)==True:
        wb2 = load_workbook(name2)
        wb=wb2
    else:
        wb = openpyxl.Workbook()

    sheet = wb.create_sheet(sheet_n)
    if in1==1:
      c1 = sheet.cell(1,1)
      c1.value=A[0]
      
      c2 = sheet.cell(1,2)
      c2.value=A[1]
      
      wb.save(name)
    else:
        a=np.size(A,axis=0)
        b=np.size(A,axis=1)
        
      
        for i in range(a):
           for j in range(b):
              c1 = sheet.cell(row=i+1, column=j+1)
              c1.value=A[i,j]
        wb.save(name)


csvFile = pandas.read_csv(r"C:\Users\user\OneDrive\Desktop\data.csv")
var1=np.array(csvFile['Invoice number'])
var2=np.array(csvFile['Date'])
var3=np.array(csvFile['customer name'])
var4=np.array(csvFile['item name'])
var5=np.array(csvFile['Qty sold'])
var6=np.array(csvFile['Unit price/USD'])
var7=np.array(csvFile['profit margin/ %'])
var8=np.array(csvFile['Total sales/USD'])
var9=np.array(csvFile['Payment recived/USD'])
var10=np.array(csvFile['Current stock'])

print("-----------total sales------------")
var11=var5*var6
print(var11)
x=sortbylist(var1,var11)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'totalsales',0)


print("-----------Qts in stock------------")
var12=var10-var5
print(var12)
x=sortbylist(var1,var12)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'stock',0)


print("-----------payment to be completed------------")
var13=var11-var9
print(var13)
x=sortbylist(var1,var13)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'due payment',0)


print("-----------total amount to be paid------------")
var15=sum(var13)

amm=[]
amm="total due amount"
amm=[amm, var15]
print(amm)
fill_array(amm,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'total amount to be paid',1)


print("-----------total amount to be paid by customer------------")
print(sortbylist(var3,var13))
x=sortbylist(var3,var13)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'customer due payment',0)


print("-----------total sales by customer------------")
print(sortbylist(var3,var11))
x=sortbylist(var3,var11)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'sales per customer',0)


print("-----------total sales by item number------------")
print(sortbylist(var4,var11))
x=sortbylist(var4,var11)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'sales per item',0)


print("-----------total sales by date------------")
print(sortbylist(var2,var11))
x=sortbylist(var2,var11)
fill_array(x,r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'sales per day',0)

print("_______income statement_______________________")
rent=np.array(csvFile['Rent'])
salery=np.array(csvFile['Salaries'])
util=np.array(csvFile['Utilities'])
misexp=np.array(csvFile['Miscellaneous Expenses'])

total_exp=rent+salery+util+misexp
print("-----------total expenses--------")
print(total_exp[0])
print("----total revenue-------")
total_sales=sum(var11)
print(total_sales)
print("-----net profit------")
profit=total_sales-total_exp[0]
print(profit)

aml=["total expenses","total revenue","net profit"]
amn=[total_exp[0],total_sales, profit]
amm=[aml, amn]
print(amm)
fill_array(np.array(amm),r"C:\Users\user\OneDrive\Desktop\Q1.xlsx",'income-statement',0)

